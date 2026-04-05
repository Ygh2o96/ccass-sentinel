#!/usr/bin/env python3
"""CCASS Sentinel — New Listing Auto-Discovery v2.0

REWRITE: Scrapes HKEX ListOfSecurities.xlsx instead of brute-forcing CCASS.
One HTTP call (~1.3 MB Excel) → parse 2,700+ equities → diff against watchlist.
Runtime: <10 seconds vs 60+ minutes.

Source: https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx
Updated daily by HKEX. Contains all listed securities with category, sub-category, ISIN, CCASS eligibility.
"""

import json, os, sys, io, urllib.request
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
WATCHLIST_FILE = REPO_ROOT / "data" / "watchlist.json"
SNAPSHOT_FILE = REPO_ROOT / "data" / "hkex_securities_snapshot.json"
HKEX_URL = "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx"

TG_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TG_CHAT = os.environ.get("TELEGRAM_CHAT_ID", "")

# Which categories are equity-like (worth monitoring in CCASS)
EQUITY_CATEGORIES = {"Equity", "Real Estate Investment Trusts", "Stapled Securities"}
# Sub-categories to include
EQUITY_SUBCATEGORIES_INCLUDE = {
    "Equity Securities (Main Board)",
    "Equity Securities (GEM)",
    "Real Estate Investment Trusts (Main Board)",
}


def tg(text):
    if not TG_TOKEN or not TG_CHAT:
        return print(f"[TG skip] {text[:80]}")
    try:
        d = json.dumps({"chat_id": int(TG_CHAT), "text": text}).encode()
        urllib.request.urlopen(urllib.request.Request(
            f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage", d,
            headers={"Content-Type": "application/json"}), timeout=10)
    except Exception as e:
        print(f"[TG err] {e}")


def download_securities_list():
    """Download and parse HKEX ListOfSecurities.xlsx → dict of equity codes."""
    print("  Downloading ListOfSecurities.xlsx...")
    req = urllib.request.Request(HKEX_URL, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    data = urllib.request.urlopen(req, timeout=60).read()
    print(f"  Downloaded: {len(data)/1024:.0f} KB")

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=False)
    ws = wb.active

    # Parse header row (row 3) to find column indices
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=c).value
        if val:
            headers[str(val).strip()] = c

    col_code = headers.get("Stock Code", 1)
    col_name = headers.get("Name of Securities", 2)
    col_cat = headers.get("Category", 3)
    col_subcat = headers.get("Sub-Category", 4)
    col_ccass = headers.get("Admitted to CCASS", None)

    # Extract update date from row 2
    update_info = str(ws.cell(row=2, column=1).value or "")
    print(f"  HKEX data: {update_info}")

    equities = {}
    for r in range(4, ws.max_row + 1):
        code = ws.cell(row=r, column=col_code).value
        name = ws.cell(row=r, column=col_name).value
        cat = ws.cell(row=r, column=col_cat).value
        subcat = ws.cell(row=r, column=col_subcat).value
        ccass = ws.cell(row=r, column=col_ccass).value if col_ccass else None

        if not code or not cat:
            continue

        code = str(code).strip().zfill(5)
        cat = str(cat).strip()

        # Filter: equity-like securities only
        if cat not in EQUITY_CATEGORIES:
            continue

        # Optional: further filter by sub-category
        subcat_str = str(subcat).strip() if subcat else ""

        equities[code] = {
            "name": str(name).strip() if name else "",
            "category": cat,
            "sub_category": subcat_str,
            "ccass": str(ccass).strip() if ccass else "",
        }

    wb.close()
    print(f"  Parsed {len(equities)} equity securities")
    return equities, update_info


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    print(f"🔍 CCASS Sentinel — New Listing Discovery v2.0 ({today})")
    print(f"  Method: HKEX ListOfSecurities.xlsx (single HTTP call)")

    # Download current securities list
    try:
        hkex_equities, update_info = download_securities_list()
    except Exception as e:
        msg = f"❌ Failed to download HKEX securities list: {e}"
        print(msg)
        tg(msg)
        sys.exit(1)

    # Load current watchlist
    wl = json.loads(WATCHLIST_FILE.read_text())
    existing_codes = {w["code"] for w in wl}
    print(f"  Current watchlist: {len(existing_codes)} stocks")

    # Load previous snapshot (if exists) for change detection
    prev_snapshot = {}
    if SNAPSHOT_FILE.exists():
        prev_snapshot = json.loads(SNAPSHOT_FILE.read_text())
        print(f"  Previous snapshot: {len(prev_snapshot)} equities")

    # Find new listings: in HKEX but NOT in watchlist
    new_codes = sorted(set(hkex_equities.keys()) - existing_codes)

    # If we have a previous snapshot, also identify truly new listings
    # (appeared in HKEX since last run)
    if prev_snapshot:
        truly_new = sorted(set(hkex_equities.keys()) - set(prev_snapshot.keys()))
        if truly_new:
            print(f"\n  🆕 Truly new listings since last scan: {len(truly_new)}")
            for code in truly_new:
                info = hkex_equities[code]
                print(f"    {code} | {info['name']:30s} | {info['sub_category']}")

    # Also detect delistings
    if prev_snapshot:
        delisted = sorted(set(prev_snapshot.keys()) - set(hkex_equities.keys()))
        if delisted:
            print(f"\n  ⚠️ Delisted since last scan: {len(delisted)}")
            for code in delisted[:10]:
                print(f"    {code} | {prev_snapshot[code].get('name', '?')}")

    # Save current snapshot
    SNAPSHOT_FILE.write_text(json.dumps(hkex_equities, indent=2, ensure_ascii=False))
    print(f"  💾 Snapshot saved: {len(hkex_equities)} equities")

    # Add new codes to watchlist
    if not new_codes:
        print(f"\n  ✅ No new listings to add (watchlist covers all {len(existing_codes)} equities)")
        # Still push Telegram summary
        if prev_snapshot and truly_new:
            tg(f"🔍 Discovery — {today}\n"
               f"HKEX: {len(hkex_equities)} equities\n"
               f"New since last scan: {len(truly_new)}\n"
               f"Already in watchlist: all covered")
        return

    added = []
    for code in new_codes:
        info = hkex_equities[code]
        # Determine tier: new IPOs (not in prev snapshot) get HIGH, old misses get MEDIUM
        is_fresh = code not in prev_snapshot if prev_snapshot else True
        tier = "HIGH" if is_fresh else "MEDIUM"

        wl.append({
            "code": code,
            "name": info["name"],
            "tier": tier,
            "source": f"AUTO-DISCOVERED {today} (HKEX ListOfSecurities)",
            "sub_category": info["sub_category"],
        })
        added.append((code, info["name"], tier))
        print(f"  ➕ {code} | {info['name']:30s} | {tier}")

    wl.sort(key=lambda x: x["code"])
    WATCHLIST_FILE.write_text(json.dumps(wl, indent=2, ensure_ascii=False))
    print(f"\n  💾 Watchlist updated: {len(wl)} stocks (+{len(added)} new)")

    # Telegram
    msg_lines = [
        f"🔍 Discovery — {today}",
        f"HKEX equities: {len(hkex_equities)}",
        f"Added to watchlist: {len(added)}",
    ]
    for code, name, tier in added[:15]:
        msg_lines.append(f"  {code} {name} [{tier}]")
    if len(added) > 15:
        msg_lines.append(f"  ... +{len(added)-15} more")
    tg("\n".join(msg_lines))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"\n  💥 FATAL ERROR:\n{tb}")
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            from telegram_push import push_error
            push_error(datetime.now().strftime("%Y-%m-%d"),
                       f"discover_new_listings.py crashed:\n{str(e)[:200]}")
        except:
            pass
        sys.exit(1)
